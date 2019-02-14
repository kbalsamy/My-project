import asyncio
from selenium import webdriver
import selenium
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
import os
import sqlite3
from xlsxwriter import Workbook


# take input from locale
url = os.environ.get("login_url")
pword = os.environ.get("p_word")


month_dict = {'January': '#mat-option-0 > .mat-option-text', 'February': '#mat-option-1 > .mat-option-text',
              'March': '#mat-option-2 > .mat-option-text', 'April': '#mat-option-3 > .mat-option-text',
              'May': '#mat-option-4 > .mat-option-text', 'June': '#mat-option-5 > .mat-option-text',
              'July': '#mat-option-6 > .mat-option-text', 'August': '#mat-option-7 > .mat-option-text',
              'September': '#mat-option-8 > .mat-option-text', 'October': '#mat-option-9 > .mat-option-text',
              'November': '#mat-option-10 > .mat-option-text', 'December': '#mat-option-11 > .mat-option-text'}


def getID():
    with open('consumerID.txt', 'r') as file:
        return [ID.strip() for ID in file]


def db_connect():
    con = sqlite3.connect('ReadingsV1onAPP.db')
    return con


async def scrape(loop, x, y, uname, pword):
    " webdriver created herewith "
    options = Options()
    options.headless = True
    binary = FirefoxBinary("C:/Program Files/Firefox Nightly/firefox.exe")
    driver = webdriver.Firefox(options=options, firefox_binary=binary)
    if not uname:
        tab = 0
        for CID in getID()[x:y]:
            driver.execute_script("window.open('about:blank', 'tab{}');".format(tab))
            driver.switch_to.window("tab{}".format(tab))
            await get_values(loop, driver, url, CID, pword, 'January', 2019)
            tab += 1
        driver.quit()
    else:
        await get_values(loop, driver, url, uname, pword, 'January', 2019)


async def get_values(loop, driver, url, ID, pword, month, year):
    "Scaping logic "
    try:
        browser = driver.get(url)
        login = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mat-input-0")))
        login.send_keys(ID)
        password = driver.find_element_by_id("mat-input-1")
        password.send_keys(pword)
        submit = driver.find_element_by_css_selector(".mat-raised-button").click()
        try:
            reading = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".mat-list-item:nth-child(2) span:nth-child(1)")))
            reading.click()
            menu_selector = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#mat-select-1 .mat-select-arrow"))).click()
            element = driver.find_element_by_css_selector(".ng-trigger-transformPanel")
            actions = ActionChains(driver)
            actions.move_to_element(element).perform()
            select_month = driver.find_element_by_css_selector(month_dict.get(month)).click()

            enter_year = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "mat-input-5")))
            enter_year.clear()
            enter_year.send_keys(year)
            try:
                enter_button = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".primary > .mat-button-wrapper"))).click()

                fetch_results = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".mr-1 > .mat-button-wrapper"))).click()
                soup = BS(driver.page_source, "lxml")
                html = soup.prettify()
                status = "Successfully extracted the values for {}".format(ID)
                await extract_values(ID, html, db_connect(), month, year, status)

                logout_menu = driver.find_element_by_css_selector(".ml-xs .mat-icon").click()
                logout = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR,
                                                                                         ".mat-menu-item:nth-child(5) > .mat-menu-ripple")))
                logout.click()
            except selenium.common.exceptions.TimeoutException as e:
                status = "Readings are not found for {}.".format(ID)
                await extract_values(ID, None, db_connect(), month, year, status)
                # driver.close()

        except selenium.common.exceptions.TimeoutException as e:
            status = "consumer {} not found.".format(ID)
            await extract_values(ID, None, db_connect(), month, year, status)
            # driver.close()

    except (selenium.common.exceptions.WebDriverException, selenium.common.exceptions.TimeoutException) as e:
        print("Check your network settings")
        driver.quit()
        loop.stop()


async def extract_values(ID, html, con, month, year, status):
    "processing the scrapped values"
    CID = ID
    tablename = month + str(year)
    if html:
        print('extracting values for {}'.format(ID))
        soup = BS(html, "html.parser")
        tab = soup.find("ngx-datatable")
        headers = tab.find_all("span")
        t_list = []
        f_list = []
        for i in headers:
            t_list.append(i.text.strip())
        for i in t_list:
            if i not in f_list and i != "":
                f_list.append(i)
        val = tab.find_all("datatable-body-cell")
        val_list = []
        for i in val:
            val_list.append(i.text.strip())
        results = [val_list[i:i + 9] for i in range(0, len(val_list), 9)]
        cursor = con.cursor()
        create_table = """CREATE TABLE IF NOT EXISTS {} (ID INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, consumer TEXT, Connection TEXT, ImportUnits REAL, ExportUnits REAL, DifUnits REAL)""".format(tablename)
        cursor.execute(create_table)
        for item in results:
            differnce = float(item[8]) - float(item[4])
            insert_values = """INSERT INTO {} (consumer, Connection, ImportUnits, ExportUnits, DifUnits) VALUES(?, ?, ?, ?, ?)""".format(tablename)
            cursor.execute(insert_values, (CID, item[0], float(item[4]), float(item[8]), differnce))
            print('values updated into database for {}'.format(CID))
            con.commit()
    else:
        print(status)
        with open('failed summary.txt', 'a+') as file:
            file.write("{} failed to get values\n".format(CID))
            file.close()


class Downloader():

    def _init_(self, tablename, filename, db_file):
        self.tablename = tablename
        self.filename = filename

    def create_sheet(self, tab='sheet1'):
        workbook = Workbook('{}.xlsx'.format(self.filename))
        worksheet = workbook.add_worksheet(tab)
        return worksheet

    def write_excel(self, row, column, db_results, worksheet):
        r = row
        c = column
        for result in db_results:
            imp, exp, dif = (result)
            worksheet.write(r, c, imp)
            worksheet.write(r, c + 5, exp)
            worksheet.write(r, c + 10, dif)
            r += 1

    def write_cons_excel(self, row, column, db_results, worksheet):
        r = row
        c = column
        previous_consumer = None
        for cons in db_results:
            consumer, n = (cons)
            if consumer != previous_consumer or consumer == None:
                worksheet.write(r, c, consumer)
                r += 1
                previous_consumer = consumer

            else:
                pass

    def get_table_name(self):
        cursor = db_connect()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        table_list = cursor.fetchall()
        t, u = (table_list)
        table_name = t[0]
        return table_name

    def get_consumer_list(self):
        table_name = self.get_table_name()
        cursor = db_connect()
        cursor.execute(("SELECT consumer, Connection FROM ?)", (table_name)))
        cons = cursor.fetchall()
        return cons

    def get_queries_from_db(self):
        table_name = self.get_table_name()
        cursor = db_connect()
        table_name = result_container = []

        cursor.execute(("SELECT ImportUnits, ExportUnits, DifUnits FROM ? WHERE  connection = 'C1'"), (table_name))
        c_1 = cursor.fetchall()
        result_container.append(c_1)

        cursor.execute(("SELECT ImportUnits, ExportUnits, DifUnits FROM ? WHERE  connection = 'C2'"), (table_name))
        c_2 = cursor.fetchall()
        result_container.append(c_2)

        cursor.execute(("SELECT ImportUnits, ExportUnits, DifUnits FROM ? WHERE  connection = 'C3'"), (table_name))
        c_3 = cursor.fetchall()
        result_container.append(c_3)

        cursor.execute(("SELECT ImportUnits, ExportUnits, DifUnits FROM ? WHERE  connection = 'C4'"), (table_name))
        c_4 = cursor.fetchall()
        result_container.append(c_4)

        cursor.execute(("SELECT ImportUnits, ExportUnits, DifUnits FROM ? WHERE  connection = 'C5'"), (table_name))
        c_5 = cursor.fetchall()
        result_container.append(c_5)

        return result_container

    def writer(self):
        cons = self.get_consumer_list()
        results = self.get_queries_from_db()
        worksheet = self.create_sheet()
        self.write_cons_excel(0, 0, cons, worksheet)
        c = 1
        for i in results:
            self.write_excel(0, c, i, worksheet)
            c += 1

    def save(self):
        workbook.close()

#  task is created


async def main(loop, i, j, uname, pword):
    "tasks created"
    await asyncio.wait([scrape(loop, i, j, uname, pword)])


def asyn_loop(x, y, uname, pword):
    "main entry "
    i = int(x)
    j = int(y)
    loop = asyncio.get_event_loop()
    loop.run_until_complete(main(loop, i, j, uname, pword))
    loop.stop()
